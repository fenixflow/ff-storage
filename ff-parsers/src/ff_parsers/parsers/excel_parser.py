"""
Excel parser implementation using openpyxl.
"""

from pathlib import Path
from typing import Optional, Union, List
import openpyxl
from openpyxl.utils import get_column_letter
import csv
import chardet

from ..base import BaseParser, ParseOptions
from ..models import ExtractedDocument, ExtractedText, ExtractedTable, DocumentMetadata
from ..exceptions import ExtractionError, CorruptedFileError
from ..utils.cleaning import clean_text, remove_extra_spaces_in_table


class ExcelParser(BaseParser):
    """Parser for Excel spreadsheets and CSV files."""

    def get_supported_extensions(self) -> list[str]:
        """Get supported file extensions."""
        return [".xlsx", ".xlsm", ".xls", ".csv", ".XLSX", ".XLSM", ".XLS", ".CSV"]

    def get_mime_types(self) -> list[str]:
        """Get supported MIME types."""
        return [
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "application/vnd.ms-excel",
            "text/csv",
            "application/csv",
        ]

    def validate(self, file_path: Union[str, Path]) -> bool:
        """Validate that the file is a valid Excel/CSV file."""
        try:
            path = Path(file_path)
            if not path.exists() or not path.is_file():
                return False

            # Check extension
            if path.suffix.lower() not in [".xlsx", ".xlsm", ".xls", ".csv"]:
                return False

            if path.suffix.lower() == ".csv":
                # Try to read first line of CSV
                with open(path, "rb") as f:
                    f.read(1024)  # Just check we can read it
                return True
            else:
                # Try to open with openpyxl
                wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
                wb.close()
                return True

        except Exception:
            return False

    def extract_metadata(self, file_path: Union[str, Path]) -> DocumentMetadata:
        """Extract Excel metadata."""
        path = self._ensure_path(file_path)
        metadata = DocumentMetadata()

        try:
            metadata.file_size = self._get_file_size(path)

            if path.suffix.lower() == ".csv":
                metadata.mime_type = "text/csv"
                # CSV files don't have built-in metadata
            else:
                metadata.mime_type = (
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

                # Load workbook for metadata
                wb = openpyxl.load_workbook(path, read_only=True)

                # Get properties
                props = wb.properties
                if props:
                    metadata.title = props.title
                    metadata.author = props.creator
                    metadata.subject = props.subject
                    metadata.created_date = props.created
                    metadata.modified_date = props.modified

                    if props.keywords:
                        metadata.keywords = [k.strip() for k in props.keywords.split(",")]

                    # Custom properties
                    metadata.custom_properties["last_modified_by"] = props.lastModifiedBy
                    metadata.custom_properties["sheet_count"] = len(wb.sheetnames)

                wb.close()

        except Exception as e:
            self._log_warning(f"Failed to extract metadata: {e}")

        return metadata

    def parse(
        self, file_path: Union[str, Path], options: Optional[ParseOptions] = None
    ) -> ExtractedDocument:
        """Parse Excel/CSV file and extract content."""
        path = self._ensure_path(file_path)
        options = options or self._get_default_options()

        # Create base document
        document = self._create_base_document(path)

        try:
            if path.suffix.lower() == ".csv":
                self._parse_csv(path, document, options)
            else:
                self._parse_excel(path, document, options)

            # Extract metadata
            if options.extract_metadata:
                document.metadata = self.extract_metadata(path)

            # Count words
            if document.text:
                document.metadata.word_count = len(document.text.split())

        except Exception as e:
            raise CorruptedFileError(
                f"Failed to parse Excel file: {str(e)}", partial_result=document
            )

        return document

    def _parse_excel(self, path: Path, document: ExtractedDocument, options: ParseOptions):
        """Parse Excel file using openpyxl."""
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

        try:
            all_text = []

            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]

                # Extract sheet content as text
                sheet_text = self._extract_sheet_text(sheet, sheet_name, options)
                if sheet_text:
                    document.pages.append(sheet_text)
                    all_text.append(sheet_text.content)

                # Extract tables
                if options.extract_tables:
                    sheet_tables = self._extract_sheet_tables(sheet, sheet_name)
                    document.tables.extend(sheet_tables)

            # Combine all text
            document.text = "\n\n".join(all_text)

        finally:
            wb.close()

    def _parse_csv(self, path: Path, document: ExtractedDocument, options: ParseOptions):
        """Parse CSV file."""
        # Detect encoding
        encoding = self._detect_csv_encoding(path)

        try:
            with open(path, "r", encoding=encoding) as f:
                # Detect delimiter
                sample = f.read(4096)
                f.seek(0)
                sniffer = csv.Sniffer()
                delimiter = sniffer.sniff(sample).delimiter

                # Read CSV
                reader = csv.reader(f, delimiter=delimiter)
                rows = list(reader)

                if not rows:
                    return

                # Create table
                if options.extract_tables and len(rows) > 1:
                    table = ExtractedTable(headers=rows[0], rows=rows[1:], name="CSV_Data")
                    document.tables.append(table)

                # Convert to text
                text_lines = []
                for row in rows:
                    text_lines.append(delimiter.join(str(cell) for cell in row))

                text = "\n".join(text_lines)

                document.text = text
                document.pages.append(ExtractedText(content=text, confidence=1.0))

        except Exception as e:
            raise ExtractionError(f"Failed to parse CSV: {str(e)}")

    def _detect_csv_encoding(self, path: Path) -> str:
        """Detect encoding of CSV file."""
        try:
            with open(path, "rb") as f:
                raw_data = f.read(10000)
                result = chardet.detect(raw_data)
                encoding = result["encoding"]

                if not encoding:
                    encoding = "utf-8"

                return encoding
        except Exception:
            return "utf-8"

    def _extract_sheet_text(
        self, sheet, sheet_name: str, options: ParseOptions
    ) -> Optional[ExtractedText]:
        """Extract text from an Excel sheet."""
        try:
            text_parts = []

            # Iterate through rows
            for row in sheet.iter_rows(values_only=True):
                # Skip empty rows
                if all(cell is None for cell in row):
                    continue

                # Convert row to text
                row_text = "\t".join(str(cell) if cell is not None else "" for cell in row)
                text_parts.append(row_text)

            if not text_parts:
                return None

            text = "\n".join(text_parts)

            # Clean text if requested
            if not options.preserve_whitespace:
                text = clean_text(text, preserve_paragraphs=False)

            return ExtractedText(content=text, sheet=sheet_name, confidence=1.0)

        except Exception as e:
            self._log_warning(f"Failed to extract text from sheet {sheet_name}: {e}")
            return None

    def _extract_sheet_tables(self, sheet, sheet_name: str) -> List[ExtractedTable]:
        """Extract tables from an Excel sheet."""
        tables = []

        try:
            # Check for Excel tables (ListObjects)
            if hasattr(sheet, "tables") and sheet.tables:
                for table_name, table_ref in sheet.tables.items():
                    table = self._extract_excel_table(sheet, table_name, table_ref, sheet_name)
                    if table:
                        tables.append(table)
            else:
                # Extract the entire sheet as a table if it has data
                table = self._extract_sheet_as_table(sheet, sheet_name)
                if table:
                    tables.append(table)

        except Exception as e:
            self._log_warning(f"Failed to extract tables from sheet {sheet_name}: {e}")

        return tables

    def _extract_excel_table(
        self, sheet, table_name: str, table_ref: str, sheet_name: str
    ) -> Optional[ExtractedTable]:
        """Extract a named Excel table."""
        try:
            # Parse table reference (e.g., "A1:D10")
            min_col, min_row, max_col, max_row = self._parse_cell_range(table_ref)

            # Extract headers
            headers = []
            for col in range(min_col, max_col + 1):
                cell_value = sheet.cell(row=min_row, column=col).value
                headers.append(str(cell_value) if cell_value is not None else "")

            # Extract rows
            rows = []
            for row in range(min_row + 1, max_row + 1):
                row_data = []
                for col in range(min_col, max_col + 1):
                    cell_value = sheet.cell(row=row, column=col).value
                    row_data.append(str(cell_value) if cell_value is not None else "")
                rows.append(row_data)

            # Clean table data
            rows = remove_extra_spaces_in_table(rows)

            return ExtractedTable(
                headers=headers, rows=rows, sheet=sheet_name, name=table_name, cell_range=table_ref
            )

        except Exception as e:
            self._log_warning(f"Failed to extract table {table_name}: {e}")
            return None

    def _extract_sheet_as_table(self, sheet, sheet_name: str) -> Optional[ExtractedTable]:
        """Extract entire sheet as a table."""
        try:
            # Get all data
            data = []
            for row in sheet.iter_rows(values_only=True):
                # Skip completely empty rows
                if all(cell is None for cell in row):
                    continue
                row_data = [str(cell) if cell is not None else "" for cell in row]
                data.append(row_data)

            if len(data) < 2:
                return None

            # Clean table data
            data = remove_extra_spaces_in_table(data)

            # First row as headers
            headers = data[0]
            rows = data[1:]

            # Determine cell range
            max_row = sheet.max_row
            max_col = sheet.max_column
            cell_range = f"A1:{get_column_letter(max_col)}{max_row}"

            return ExtractedTable(
                headers=headers,
                rows=rows,
                sheet=sheet_name,
                name=f"{sheet_name}_Table",
                cell_range=cell_range,
            )

        except Exception as e:
            self._log_warning(f"Failed to extract sheet as table: {e}")
            return None

    def _parse_cell_range(self, range_str: str) -> tuple:
        """Parse cell range string (e.g., 'A1:D10') to column and row numbers."""
        try:
            parts = range_str.split(":")
            if len(parts) != 2:
                raise ValueError("Invalid range format")

            # Parse start cell
            start_cell = parts[0]
            start_col = openpyxl.utils.column_index_from_string(
                "".join(c for c in start_cell if c.isalpha())
            )
            start_row = int("".join(c for c in start_cell if c.isdigit()))

            # Parse end cell
            end_cell = parts[1]
            end_col = openpyxl.utils.column_index_from_string(
                "".join(c for c in end_cell if c.isalpha())
            )
            end_row = int("".join(c for c in end_cell if c.isdigit()))

            return start_col, start_row, end_col, end_row

        except Exception as e:
            raise ValueError(f"Failed to parse cell range '{range_str}': {e}")
